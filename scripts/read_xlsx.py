import openpyxl
from openpyxl import load_workbook

wb = load_workbook('/home/z/my-project/upload/Porto 2026.xlsx', data_only=False)
print("Sheets:", wb.sheetnames)
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== Sheet: {sheet_name} ===")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    print()
    # Print first 30 rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(40, ws.max_row), values_only=True), 1):
        print(f"R{row_idx}: {row}")
    print()
    print("---")
    print()
